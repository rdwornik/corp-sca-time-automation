"""
Excel writer with formatting, colors and table.
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def write_excel_with_formatting(df: pd.DataFrame, output_path: Path) -> None:
    """Write DataFrame to Excel with colors and table formatting."""

    # Save basic Excel first
    df.to_excel(output_path, index=False, sheet_name="Time Entries")

    # Load and format
    wb = load_workbook(output_path)
    ws = wb.active

    # Define colors
    green_fill = PatternFill(
        start_color="90EE90", end_color="90EE90", fill_type="solid"
    )  # Original
    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )  # Autofilled
    red_fill = PatternFill(
        start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"
    )  # Week total

    # Find column indices
    columns = {cell.value: cell.column for cell in ws[1]}
    category_col = columns.get("category")
    autofill_col = columns.get("is_autofilled")
    hours_col = columns.get("hours")
    hours_col_letter = get_column_letter(hours_col) if hours_col else None

    # Track first/last data row per week for SUM formula ranges
    week_data_start: int | None = None
    week_data_end: int | None = None

    # Apply colors to data rows
    for row_idx in range(2, ws.max_row + 1):
        category_value = (
            ws.cell(row=row_idx, column=category_col).value if category_col else None
        )
        is_autofilled = (
            ws.cell(row=row_idx, column=autofill_col).value if autofill_col else False
        )

        # Determine fill color
        if category_value == ">>> WEEK TOTAL":
            fill = red_fill
            # Replace static Hours value with SUM formula over this week's data rows
            if hours_col_letter and week_data_start is not None and week_data_end is not None:
                formula = f"=SUM({hours_col_letter}{week_data_start}:{hours_col_letter}{week_data_end})"
                ws.cell(row=row_idx, column=hours_col).value = formula
            # Reset trackers for next week
            week_data_start = None
            week_data_end = None
        else:
            # Track data rows for current week
            if week_data_start is None:
                week_data_start = row_idx
            week_data_end = row_idx

            if is_autofilled:
                fill = yellow_fill
            else:
                fill = green_fill

        # Apply to all cells in row
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    # Create table
    table_range = f"A1:{chr(64 + ws.max_column)}{ws.max_row}"
    table = Table(displayName="TimeEntries", ref=table_range)

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Auto-adjust column widths
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    wb.save(output_path)
