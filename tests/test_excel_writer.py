"""Tests for Excel SUM formula generation in write_excel_with_formatting."""

import pytest
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.excel_writer import write_excel_with_formatting


def _make_df_single_week() -> pd.DataFrame:
    """Single week with 2 data rows + 1 WEEK TOTAL row."""
    return pd.DataFrame([
        {"week_beginning": "2025-03-23", "category": "Admin", "hours": 8.0, "is_autofilled": False, "comments": None, "opportunity_id": None, "client": None},
        {"week_beginning": "2025-03-23", "category": "Internal Meeting", "hours": 4.0, "is_autofilled": False, "comments": None, "opportunity_id": None, "client": None},
        {"week_beginning": "2025-03-23", "category": ">>> WEEK TOTAL", "hours": 12.0, "is_autofilled": False, "comments": None, "opportunity_id": None, "client": None},
    ])


def _make_df_single_entry_week() -> pd.DataFrame:
    """Single week with only 1 data row + WEEK TOTAL."""
    return pd.DataFrame([
        {"week_beginning": "2025-03-23", "category": "Admin", "hours": 8.0, "is_autofilled": False, "comments": None, "opportunity_id": None, "client": None},
        {"week_beginning": "2025-03-23", "category": ">>> WEEK TOTAL", "hours": 8.0, "is_autofilled": False, "comments": None, "opportunity_id": None, "client": None},
    ])


def _make_df_two_weeks() -> pd.DataFrame:
    """Two weeks with different sizes to verify independent formula ranges."""
    return pd.DataFrame([
        {"week_beginning": "2025-03-23", "category": "Admin", "hours": 8.0, "is_autofilled": False, "comments": None, "opportunity_id": None, "client": None},
        {"week_beginning": "2025-03-23", "category": "Internal Meeting", "hours": 4.0, "is_autofilled": False, "comments": None, "opportunity_id": None, "client": None},
        {"week_beginning": "2025-03-23", "category": ">>> WEEK TOTAL", "hours": 12.0, "is_autofilled": False, "comments": None, "opportunity_id": None, "client": None},
        {"week_beginning": "2025-03-30", "category": "Discovery", "hours": 6.0, "is_autofilled": False, "comments": None, "opportunity_id": None, "client": None},
        {"week_beginning": "2025-03-30", "category": ">>> WEEK TOTAL", "hours": 6.0, "is_autofilled": False, "comments": None, "opportunity_id": None, "client": None},
    ])


def _get_hours_col_letter(ws) -> str:
    """Return the column letter for the 'hours' column."""
    for cell in ws[1]:
        if cell.value == "hours":
            return get_column_letter(cell.column)
    raise ValueError("hours column not found")


def _get_category_col(ws) -> int:
    """Return the column index for the 'category' column."""
    for cell in ws[1]:
        if cell.value == "category":
            return cell.column
    raise ValueError("category column not found")


class TestWeekTotalSumFormula:
    def test_week_total_hours_is_formula(self, tmp_path):
        """WEEK TOTAL Hours cell must contain a =SUM formula, not a static float."""
        path = tmp_path / "out.xlsx"
        write_excel_with_formatting(_make_df_single_week(), path)

        wb = load_workbook(path)
        ws = wb.active
        hours_col_letter = _get_hours_col_letter(ws)
        category_col = _get_category_col(ws)

        from openpyxl.utils import column_index_from_string
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=category_col).value == ">>> WEEK TOTAL":
                hours_val = ws.cell(row=row_idx, column=column_index_from_string(hours_col_letter)).value
                assert isinstance(hours_val, str), f"Expected formula string, got {type(hours_val)}: {hours_val}"
                assert hours_val.startswith("=SUM("), f"Expected =SUM formula, got: {hours_val}"
                return
        pytest.fail("No WEEK TOTAL row found")

    def test_week_total_formula_references_correct_rows(self, tmp_path):
        """SUM range must span exactly the data rows for that week (not TOTAL row itself)."""
        path = tmp_path / "out.xlsx"
        write_excel_with_formatting(_make_df_single_week(), path)

        wb = load_workbook(path)
        ws = wb.active
        hours_col_letter = _get_hours_col_letter(ws)
        category_col = _get_category_col(ws)

        # Data rows: rows 2 and 3 (Admin and Internal Meeting)
        # WEEK TOTAL: row 4
        # Expected formula: =SUM(D2:D3) (or whatever column D is)
        from openpyxl.utils import column_index_from_string
        hours_col_idx = column_index_from_string(hours_col_letter)

        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=category_col).value == ">>> WEEK TOTAL":
                formula = ws.cell(row=row_idx, column=hours_col_idx).value
                expected = f"=SUM({hours_col_letter}2:{hours_col_letter}{row_idx - 1})"
                assert formula == expected, f"Expected {expected}, got {formula}"
                return
        pytest.fail("No WEEK TOTAL row found")

    def test_single_row_week_total_formula(self, tmp_path):
        """Single data row week: formula should be =SUM(Dn:Dn) (same start and end row)."""
        path = tmp_path / "out.xlsx"
        write_excel_with_formatting(_make_df_single_entry_week(), path)

        wb = load_workbook(path)
        ws = wb.active
        hours_col_letter = _get_hours_col_letter(ws)
        category_col = _get_category_col(ws)

        from openpyxl.utils import column_index_from_string
        hours_col_idx = column_index_from_string(hours_col_letter)

        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=category_col).value == ">>> WEEK TOTAL":
                formula = ws.cell(row=row_idx, column=hours_col_idx).value
                # Single data row is row 2, WEEK TOTAL is row 3
                expected = f"=SUM({hours_col_letter}2:{hours_col_letter}2)"
                assert formula == expected, f"Expected {expected}, got {formula}"
                return
        pytest.fail("No WEEK TOTAL row found")

    def test_multiple_weeks_have_independent_formulas(self, tmp_path):
        """Each week's TOTAL formula must reference only its own data rows, not other weeks."""
        path = tmp_path / "out.xlsx"
        write_excel_with_formatting(_make_df_two_weeks(), path)

        wb = load_workbook(path)
        ws = wb.active
        hours_col_letter = _get_hours_col_letter(ws)
        category_col = _get_category_col(ws)

        from openpyxl.utils import column_index_from_string
        hours_col_idx = column_index_from_string(hours_col_letter)

        total_rows = []
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=category_col).value == ">>> WEEK TOTAL":
                formula = ws.cell(row=row_idx, column=hours_col_idx).value
                total_rows.append((row_idx, formula))

        assert len(total_rows) == 2, f"Expected 2 WEEK TOTAL rows, found {len(total_rows)}"

        # Week 1: data rows 2-3, total at row 4 → =SUM(Hcol2:Hcol3)
        row1, formula1 = total_rows[0]
        assert formula1 == f"=SUM({hours_col_letter}2:{hours_col_letter}{row1 - 1})"

        # Week 2: data row 5, total at row 6 → =SUM(Hcol5:Hcol5)
        row2, formula2 = total_rows[1]
        assert formula2 == f"=SUM({hours_col_letter}{row1 + 1}:{hours_col_letter}{row2 - 1})"

        # Formulas must not overlap
        assert str(row1) not in formula2, "Week 2 formula references week 1 rows"


class TestPermissionErrorHandling:
    """Tests for PermissionError fast-fail and write-guard behavior."""

    def test_check_excel_writable_nonexistent_file(self, tmp_path):
        """Non-existent output path should pass silently (no file = no lock)."""
        from run import _check_excel_writable

        path = tmp_path / "does_not_exist.xlsx"
        # Should not raise
        _check_excel_writable(path)

    def test_check_excel_writable_unlocked_file(self, tmp_path):
        """Existing writable file should pass silently."""
        from run import _check_excel_writable

        path = tmp_path / "unlocked.xlsx"
        path.write_bytes(b"dummy")
        # Should not raise
        _check_excel_writable(path)

    def test_check_excel_writable_locked_file_exits(self, tmp_path, monkeypatch):
        """File locked by another process should print error and exit."""
        from run import _check_excel_writable

        path = tmp_path / "locked.xlsx"
        path.write_bytes(b"dummy")

        def _raise_permission(*args, **kwargs):
            raise PermissionError("File locked")

        monkeypatch.setattr(path.__class__, "open", _raise_permission)

        with pytest.raises(SystemExit) as exc_info:
            _check_excel_writable(path)
        assert exc_info.value.code == 1

    def test_write_excel_permission_error_raises_systemexit(self, tmp_path, monkeypatch):
        """write_excel_with_formatting PermissionError must surface as SystemExit."""
        import pandas as pd
        from src.excel_preview import generate_final_preview

        def _raise_permission(*args, **kwargs):
            raise PermissionError("File locked by Excel")

        monkeypatch.setattr("src.excel_writer.write_excel_with_formatting", _raise_permission)
        # Stub upstream pipeline so only the write step runs
        monkeypatch.setattr(
            "src.excel_preview.generate_aggregated_preview",
            lambda **kw: pd.DataFrame([
                {"week_beginning": "2025-03-23", "category": "Admin", "hours": 8.0,
                 "is_autofilled": False, "comments": None, "opportunity_id": None,
                 "client": None, "needs_review": False, "status": "NEW",
                 "title": "x", "external_domains": ""},
            ]),
        )

        output = tmp_path / "preview.xlsx"
        with pytest.raises(SystemExit) as exc_info:
            generate_final_preview(output_path=output, fill=False)
        assert "Cannot write Excel" in str(exc_info.value)
